"""MarkItDown integration, isolated behind one adapter.

Everything this project knows about MarkItDown lives in this file, so upstream can
be upgraded by changing one module.

Upstream assumptions recorded here (verified against markitdown 0.1.7)
---------------------------------------------------------------------
* ``MarkItDown.__init__`` unconditionally constructs ``magika.Magika()``, which
  loads an ONNX model. That costs real memory and a noticeable first-call delay,
  and it can fail outright on a memory-constrained host (observed once during
  development as ``RUNTIME_EXCEPTION: bad allocation``). The instance is therefore
  built lazily, exactly once per process, and a construction failure is reported as
  a clean tool error rather than taking the server down.
* We only ever call ``convert_local`` with an explicit ``StreamInfo`` carrying the
  extension and MIME type we detected ourselves. That matters for security: with a
  populated ``StreamInfo`` and no ``url``, MarkItDown's URL-driven converters
  (Bing SERP, YouTube, Wikipedia, RSS) cannot match, so no conversion can make a
  network request.
* ``DocumentConverterResult`` exposes only ``.text_content``. There is no page or
  slide structure in the API, so structure is recovered by parsing the emitted
  Markdown -- see the marker constants below.
* Emitted structure markers, confirmed empirically:
    - PPTX: ``<!-- Slide number: N -->`` before each slide -- including blank and
      image-only slides, so the sequence is contiguous 1..N for well-formed decks.
    - XLSX/XLS: ``## <sheet name>`` before each sheet's table.
    - DOCX/HTML/CSV: plain Markdown with no positional markers.
  These are not a documented contract, and slide *content* can contain a literal
  ``<!-- Slide number: ... -->`` string that also survives into the output
  (verified). Therefore structure is only trusted after cross-checking against
  ground truth: python-pptx's real slide count for PPTX, openpyxl's sheet names
  for XLSX. On any mismatch the parsing degrades to "one undifferentiated block"
  with a warning -- content is still returned, only the per-slide/per-sheet
  evidence and slide selection are dropped.
"""

from __future__ import annotations

import asyncio
import re
import threading
from importlib.metadata import PackageNotFoundError, version as package_version

from markitdown import (
    FileConversionException,
    MarkItDown,
    MissingDependencyException,
    StreamInfo,
    UnsupportedFormatException,
)

from ..errors import DocumentConversionFailedError, UnsupportedMediaTypeError
from ..models import (
    AnalyzeMediaRequest,
    EvidenceItem,
    EvidenceType,
    MediaCategory,
    MediaInfo,
    ProcessorResult,
)
from ..security.limits import parse_page_selection
from .base import ProcessingContext

SLIDE_MARKER_RE = re.compile(r"^<!-- Slide number:\s*(\d+)\s*-->\s*$", re.MULTILINE)
SHEET_HEADING_RE = re.compile(r"^##\s+(.+?)\s*$", re.MULTILINE)
HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$", re.MULTILINE)

_SUPPORTED = {
    MediaCategory.OFFICE,
    MediaCategory.HTML,
    MediaCategory.DATA,
    MediaCategory.EMAIL,
    MediaCategory.EBOOK,
    MediaCategory.NOTEBOOK,
}

_instance_lock = threading.Lock()
_instance: MarkItDown | None = None


def upstream_version() -> str:
    try:
        return package_version("markitdown")
    except PackageNotFoundError:  # pragma: no cover - only if installed oddly
        return "unknown"


def get_markitdown() -> MarkItDown:
    """Lazily build the process-wide MarkItDown instance.

    Plugins stay disabled: third-party plugins are arbitrary code that would run
    against untrusted documents, and enabling them silently would widen the trust
    boundary without the operator saying so.
    """
    global _instance
    if _instance is not None:
        return _instance
    with _instance_lock:
        if _instance is None:
            try:
                _instance = MarkItDown(enable_builtins=True, enable_plugins=False)
            except Exception as exc:  # noqa: BLE001 - upstream raises bare exceptions here
                raise DocumentConversionFailedError(
                    f"MarkItDown could not be initialised: {exc}",
                    hint=(
                        "MarkItDown loads the magika ONNX model at construction time. "
                        "This usually means the host is out of memory, or the "
                        "onnxruntime wheel does not match the CPU. Run "
                        "'media-context-mcp doctor' for details; plain-text and PDF "
                        "analysis do not need MarkItDown."
                    ),
                ) from exc
    return _instance


def reset_markitdown() -> None:
    """Drop the cached instance. Tests use this; nothing else should need it."""
    global _instance
    with _instance_lock:
        _instance = None


def convert_sync(path: str, extension: str, mimetype: str, filename: str) -> str:
    """Blocking conversion. Called from a worker thread."""
    converter = get_markitdown()
    stream_info = StreamInfo(
        extension=extension or None,
        mimetype=mimetype or None,
        filename=filename,
        # url is deliberately omitted: it is what activates the network converters.
    )
    try:
        result = converter.convert_local(path, stream_info=stream_info)
    except MissingDependencyException as exc:
        raise DocumentConversionFailedError(
            f"MarkItDown is missing an optional dependency for this format: {exc}",
            hint=(
                "Install the matching extra, e.g. "
                "pip install 'markitdown[docx,pptx,xlsx,xls,pdf,outlook]'."
            ),
        ) from exc
    except UnsupportedFormatException as exc:
        raise UnsupportedMediaTypeError(
            f"MarkItDown has no converter for this file: {exc}",
            hint="Convert the file to PDF, an Office format, HTML or plain text.",
        ) from exc
    except FileConversionException as exc:
        raise DocumentConversionFailedError(
            f"MarkItDown failed to convert the file: {exc}",
            hint=(
                "The file may be corrupt, password-protected, or an unexpected format "
                "for its extension. Try opening it in its native application and "
                "re-saving it."
            ),
        ) from exc
    return result.text_content or ""


def true_slide_count(path: str) -> int | None:
    """Ground-truth slide count from python-pptx. ``None`` when unavailable."""
    try:
        from pptx import Presentation

        return len(Presentation(path).slides._sldIdLst)  # noqa: SLF001 - stable internal
    except Exception:  # noqa: BLE001 - a broken deck must not break conversion
        return None


def true_sheet_names(path: str) -> list[str] | None:
    """Ground-truth sheet names from openpyxl. ``None`` when unavailable (.xls etc.)."""
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:  # noqa: BLE001
        return None


def split_slides(markdown: str, expected_count: int | None) -> list[tuple[int, str]]:
    """Split PPTX output into ``(slide_number, body)`` pairs.

    Slide content can itself contain a literal marker string, which would corrupt
    the split. The parse is therefore only trusted when the found markers form
    exactly the contiguous sequence ``1..expected_count``; anything else returns
    ``[]`` and the caller degrades to unstructured content.
    """
    matches = list(SLIDE_MARKER_RE.finditer(markdown))
    if not matches:
        return []
    numbers = [int(match.group(1)) for match in matches]
    if expected_count is not None and numbers != list(range(1, expected_count + 1)):
        return []
    if expected_count is None and numbers != list(range(1, len(numbers) + 1)):
        return []
    slides: list[tuple[int, str]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(markdown)
        slides.append((numbers[index], markdown[start:end].strip()))
    return slides


def split_sheets(markdown: str, expected_names: list[str] | None) -> list[tuple[str, str]]:
    """Split XLSX output into ``(sheet_name, body)`` pairs.

    When ground-truth names are available the parsed headings must match them
    exactly (order included); otherwise the parse is rejected and the caller
    degrades to unstructured content. Without ground truth (.xls via xlrd) the
    headings are used as-is but the caller labels them as unverified.
    """
    matches = list(SHEET_HEADING_RE.finditer(markdown))
    if not matches:
        return []
    names = [match.group(1).strip() for match in matches]
    if expected_names is not None and names != expected_names:
        return []
    sheets: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(markdown)
        sheets.append((names[index], markdown[start:end].strip()))
    return sheets


def _outline(markdown: str, limit: int = 12) -> list[str]:
    headings = [
        f"{'#' * len(match.group(1))} {match.group(2)}"
        for match in HEADING_RE.finditer(markdown)
    ]
    return headings[:limit]


def _excerpt(text: str, limit: int = 400) -> str:
    condensed = text.strip()
    if len(condensed) <= limit:
        return condensed
    return condensed[:limit].rstrip() + " ..."


class MarkItDownProcessor:
    """Converts Office, HTML, CSV, e-mail, EPUB and notebook files to Markdown."""

    name = "markitdown"
    # Bump the local part on any change that alters output for identical input. The
    # upstream MarkItDown version is folded in so that upgrading the dependency
    # invalidates exactly this processor's cache entries, automatically.
    version = f"1.0.0+markitdown-{upstream_version()}"

    def supports(self, info: MediaInfo) -> bool:
        return info.category in _SUPPORTED

    async def process(
        self,
        request: AnalyzeMediaRequest,
        context: ProcessingContext,
    ) -> ProcessorResult:
        info = context.info
        markdown = await asyncio.to_thread(
            convert_sync, str(info.path), info.extension, info.mime_type, info.name
        )

        warnings: list[str] = []
        if not markdown.strip():
            warnings.append(
                "MarkItDown produced no text for this file. It may contain only images "
                "or embedded objects. For a visual reading, export it to PDF or an "
                "image and call analyze_media again with mode='vision'."
            )

        is_pptx = info.extension == ".pptx"
        is_xlsx = info.extension == ".xlsx"

        slides: list[tuple[int, str]] = []
        sheets: list[tuple[str, str]] = []
        sheets_verified = False

        if is_pptx:
            expected = await asyncio.to_thread(true_slide_count, str(info.path))
            slides = split_slides(markdown, expected)
            if not slides:
                warnings.append(
                    "Slide boundaries could not be verified against the deck's real "
                    "slide count, so per-slide evidence and slide selection are "
                    "unavailable for this file; the full converted content is returned "
                    "as one block."
                )
        elif info.category is MediaCategory.OFFICE and info.extension in {".xlsx", ".xls"}:
            expected_names = (
                await asyncio.to_thread(true_sheet_names, str(info.path)) if is_xlsx else None
            )
            sheets = split_sheets(markdown, expected_names)
            sheets_verified = expected_names is not None and bool(sheets)
            if is_xlsx and expected_names is not None and not sheets:
                warnings.append(
                    "Sheet headings in the converted output did not match the "
                    "workbook's real sheet names, so per-sheet evidence is unavailable; "
                    "the full converted content is returned as one block."
                )
            elif sheets and not sheets_verified:
                warnings.append(
                    "Sheet names were parsed from the converted output but could not be "
                    "verified against the workbook; treat the sheet boundaries as "
                    "approximate."
                )

        evidence: list[EvidenceItem] = []
        selected_note: str | None = None

        if slides:
            markdown, evidence, selected_note = self._handle_slides(
                slides, request, context, warnings
            )
        elif sheets:
            evidence = [
                EvidenceItem(
                    type=EvidenceType.SHEET,
                    location=name,
                    content=_excerpt(body),
                )
                for name, body in sheets
            ]
            if request.pages:
                warnings.append(
                    "The 'pages' parameter does not apply to spreadsheets; all sheets "
                    "were converted. Sheet names are listed in the evidence."
                )
        else:
            evidence = [
                EvidenceItem(
                    type=EvidenceType.TEXT,
                    location=info.name,
                    content=_excerpt(markdown),
                )
            ]
            if request.pages:
                warnings.append(
                    f"The 'pages' parameter does not apply to {info.extension or 'this'} "
                    "files and was ignored."
                )

        summary = self._summarise(info, markdown, slides, sheets, selected_note)

        return ProcessorResult(
            processor=self.name,
            processor_version=self.version,
            summary=summary,
            content_markdown=markdown,
            evidence=evidence,
            warnings=warnings,
            extra={
                "slide_count": len(slides),
                "sheet_names": [name for name, _ in sheets],
                "outline": _outline(markdown),
            },
        )

    def _handle_slides(
        self,
        slides: list[tuple[int, str]],
        request: AnalyzeMediaRequest,
        context: ProcessingContext,
        warnings: list[str],
    ) -> tuple[str, list[EvidenceItem], str | None]:
        """Apply slide selection and build per-slide evidence."""
        total = len(slides)
        selected_numbers = parse_page_selection(
            request.pages, total, context.settings.max_pages
        )
        if request.pages and len(selected_numbers) < total:
            note = f"slides {request.pages} of {total}"
        elif len(selected_numbers) < total:
            note = f"first {len(selected_numbers)} of {total} slides (max_pages limit)"
            warnings.append(
                f"Only the first {len(selected_numbers)} of {total} slides were "
                f"converted because MEDIA_MCP_MAX_PAGES={context.settings.max_pages}. "
                "Use the 'pages' parameter to reach later slides."
            )
        else:
            note = None

        wanted = set(selected_numbers)
        kept = [(number, body) for number, body in slides if number in wanted]

        rebuilt = "\n\n".join(
            f"### Slide {number}\n\n{body}" if body else f"### Slide {number}\n\n_(no text)_"
            for number, body in kept
        )
        evidence = [
            EvidenceItem(
                type=EvidenceType.SLIDE,
                location=f"slide {number}",
                content=_excerpt(body) or "(no text on this slide)",
            )
            for number, body in kept
        ]
        return rebuilt, evidence, note

    def _summarise(
        self,
        info: MediaInfo,
        markdown: str,
        slides: list[tuple[int, str]],
        sheets: list[tuple[str, str]],
        selected_note: str | None,
    ) -> str:
        label = info.extension.lstrip(".").upper() or info.mime_type
        parts = [f"{label} document `{info.name}` converted to Markdown by MarkItDown"]
        if slides:
            parts.append(f"{len(slides)} slide(s)" + (f", returning {selected_note}" if selected_note else ""))
        if sheets:
            parts.append(f"sheets: {', '.join(name for name, _ in sheets)}")
        headings = _outline(markdown, limit=4)
        if headings and not sheets:
            parts.append("top headings: " + "; ".join(h.lstrip('# ') for h in headings))
        parts.append(f"{len(markdown):,} characters of Markdown")
        return ". ".join(parts) + "."
